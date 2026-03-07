"""
Output writers: write DataFrame to CSV/XLSX/JSON/YAML and write report to JSON.
"""

from pathlib import Path
from typing import Any

import pandas as pd


def write_data(
    df: pd.DataFrame,
    path: str | Path,
    format: str | None = None,
    **kwargs: Any,
) -> None:
    """
    Write DataFrame to file. format: csv | xlsx | json | yaml.
    If None, inferred from path extension.
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    fmt = (format or _infer_format(path)).lower()

    if fmt == "csv":
        df.to_csv(path, index=False, **kwargs)
    elif fmt == "xlsx":
        df.to_excel(path, index=False, engine="openpyxl", **kwargs)
        _format_xlsx(path)
    elif fmt == "json":
        df.to_json(path, orient="records", indent=2, **kwargs)
    elif fmt == "yaml":
        _write_yaml(df, path, **kwargs)
    else:
        raise ValueError(f"Unsupported output format: {format}")


def _write_yaml(df: pd.DataFrame, path: Path, **kwargs: Any) -> None:
    import yaml
    records = df.to_dict(orient="records")
    path.write_text(yaml.dump(records, default_flow_style=False, allow_unicode=True), encoding="utf-8")


def _format_xlsx(path: Path) -> None:
    """
    Apply minimal professional formatting to an XLSX file after data is written.
    Does not alter cell values; only appearance and sheet behavior.
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = load_workbook(path)
    try:
        sheet = wb.active
        if sheet is None:
            return

        # 1. Bold header row (row 1) for clear column labels
        for cell in sheet[1]:
            cell.font = Font(bold=True)

        # 2. Freeze top row so header stays visible when scrolling
        sheet.freeze_panes = "A2"

        # 3. Enable autofilter on header row so users can filter/sort in Excel
        if sheet.max_row > 0 and sheet.max_column > 0:
            ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
            sheet.auto_filter.ref = ref

        # 4. Auto-size columns to fit content (readable width, capped to avoid huge columns)
        for col_idx in range(1, sheet.max_column + 1):
            max_length = 0
            for row in range(1, sheet.max_row + 1):
                cell = sheet.cell(row=row, column=col_idx)
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            # openpyxl width ≈ character count; add padding, clamp to reasonable range
            width = min(max(max_length + 1, 10), 50)
            sheet.column_dimensions[get_column_letter(col_idx)].width = width

        wb.save(path)
    finally:
        wb.close()


def _infer_format(path: Path) -> str:
    ext = path.suffix.lower()
    mapping = {".csv": "csv", ".xlsx": "xlsx", ".json": "json", ".yaml": "yaml", ".yml": "yaml"}
    return mapping.get(ext, "csv")


def write_report(report_dict: dict[str, Any], path: str | Path) -> None:
    """Write report dict to JSON file."""
    import json
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(report_dict, indent=2), encoding="utf-8")
